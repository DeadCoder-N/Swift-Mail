import os
import re
import csv
import json
import time
import uuid
import smtplib
import threading
from flask import Flask, render_template, request, jsonify, Response
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import load_workbook

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

SMTP_CONFIGS = {
    "gmail":   {"host": "smtp.gmail.com",      "port": 587, "batch": 50, "delay": 2},
    "outlook": {"host": "smtp.office365.com",  "port": 587, "batch": 30, "delay": 3},
    "yahoo":   {"host": "smtp.mail.yahoo.com", "port": 587, "batch": 50, "delay": 2},
    "custom":  {"host": "",                     "port": 587, "batch": 100, "delay": 1},
}

# job_id -> {sent, total, failed, done, cancelled}
job_store = {}

def normalize_key(k):
    return str(k).strip().lower().replace(" ", "_").replace("-", "_")

def read_recipients(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    rows = []

    if ext == ".xlsx":
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return [], []
        first = [str(c).strip() if c else "" for c in all_rows[0]]
        has_headers = any(k in [normalize_key(c) for c in first]
                         for k in ["email", "first_name", "last_name", "name"])
        if has_headers:
            headers = [normalize_key(c) for c in first]
            for row in all_rows[1:]:
                obj = {headers[i]: str(row[i]).strip() if row[i] else "" for i in range(len(headers))}
                rows.append(obj)
        else:
            for row in all_rows:
                for cell in row:
                    if cell and "@" in str(cell):
                        rows.append({"email": str(cell).strip()})

    elif ext == ".csv":
        with open(filepath, newline="", encoding="utf-8") as f:
            sample = f.read(512)
            f.seek(0)
            has_headers = any(k in normalize_key(sample) for k in ["email", "first_name", "last_name"])
            if has_headers:
                reader = csv.DictReader(f)
                for row in reader:
                    rows.append({normalize_key(k): v.strip() for k, v in row.items()})
            else:
                for row in csv.reader(f):
                    for cell in row:
                        if "@" in cell:
                            rows.append({"email": cell.strip()})

    elif ext == ".json":
        with open(filepath, encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            for item in data:
                if isinstance(item, str) and "@" in item:
                    rows.append({"email": item.strip()})
                elif isinstance(item, dict):
                    rows.append({normalize_key(k): str(v).strip() for k, v in item.items()})
    else:
        raise ValueError("Unsupported file type")

    all_keys = set()
    for r in rows:
        all_keys.update(r.keys())

    return rows, list(all_keys)

def wrap_template(body, footer_text=""):
    footer_html = f'<p style="margin:0;font-size:12px;color:#9ca3af;">{footer_text}</p>' if footer_text else ""
    return f"""
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f3f4f6;font-family:'Segoe UI',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f3f4f6;padding:40px 0;">
    <tr><td align="center">
      <table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:12px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">
        <tr>
          <td style="padding:40px 48px;">
            <div style="font-size:15px;color:#1f2937;line-height:1.8;">
              {body}
            </div>
          </td>
        </tr>
        {"<tr><td style='padding:20px 48px 32px;border-top:1px solid #e5e7eb;text-align:center;'>" + footer_html + "</td></tr>" if footer_text else ""}
      </table>
    </td></tr>
  </table>
</body>
</html>"""

def personalize(text, recipient):
    first = recipient.get("first_name", "")
    last  = recipient.get("last_name", "")
    full  = f"{first} {last}".strip() or recipient.get("name", "")
    text  = text.replace("{{first_name}}", first)
    text  = text.replace("{{last_name}}", last)
    text  = text.replace("{{full_name}}", full)
    text  = text.replace("{{email}}", recipient.get("email", ""))
    return text

def _connect(host, port, sender_email, sender_pass):
    server = smtplib.SMTP(host, port, timeout=15)
    server.ehlo()
    server.starttls()
    server.login(sender_email, sender_pass)
    return server

def run_send_job(job_id, recipients, subject, body, provider, sender_email, sender_pass,
                 use_template, footer_text, personalize_on, custom_host, custom_port):
    job = job_store[job_id]
    config = SMTP_CONFIGS.get(provider, SMTP_CONFIGS["custom"])
    host   = custom_host if provider == "custom" else config["host"]
    port   = int(custom_port) if provider == "custom" else config["port"]
    batch  = config.get("batch", 50)
    delay  = config.get("delay", 2)

    MAX_RETRIES = 3

    try:
        server = _connect(host, port, sender_email, sender_pass)
    except Exception as e:
        job["error"] = str(e)
        job["done"]  = True
        return

    try:
        for i, r in enumerate(recipients):
            if job.get("cancelled"):
                break

            email = r.get("email", "")
            if not email:
                continue

            b = personalize(body, r) if personalize_on else body
            s = personalize(subject, r) if personalize_on else subject
            final_body = wrap_template(b, footer_text) if use_template else b

            msg = MIMEMultipart("alternative")
            msg["Subject"] = re.sub(r'<[^>]+>', '', s).strip()
            msg["From"]    = sender_email
            msg["To"]      = email
            msg.attach(MIMEText(final_body, "html"))

            for attempt in range(MAX_RETRIES):
                try:
                    server.sendmail(sender_email, email, msg.as_string())
                    job["sent"] += 1
                    break
                except smtplib.SMTPRecipientsRefused as e:
                    job["failed"].append({"email": email, "error": str(e)})
                    break
                except (smtplib.SMTPServerDisconnected, smtplib.SMTPSenderRefused):
                    try:
                        server = _connect(host, port, sender_email, sender_pass)
                    except Exception as ce:
                        if attempt == MAX_RETRIES - 1:
                            job["failed"].append({"email": email, "error": str(ce)})
                except Exception as e:
                    if attempt == MAX_RETRIES - 1:
                        job["failed"].append({"email": email, "error": str(e)})
                    else:
                        time.sleep(1)

            # batch delay to respect SMTP rate limits
            if (i + 1) % batch == 0 and not job.get("cancelled"):
                time.sleep(delay)

    finally:
        try:
            server.quit()
        except Exception:
            pass
        job["done"] = True

@app.route("/")
def index():
    return render_template("index.html", providers=list(SMTP_CONFIGS.keys()))

@app.route("/upload", methods=["POST"])
def upload():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400
    path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(path)
    try:
        recipients, fields = read_recipients(path)
        has_names = any(k in fields for k in ["first_name", "last_name", "name"])
        return jsonify({
            "recipients": recipients,
            "count": len(recipients),
            "fields": fields,
            "has_names": has_names
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/send", methods=["POST"])
def send():
    data           = request.json
    recipients     = data.get("recipients", [])
    subject        = data.get("subject", "")
    body           = data.get("body", "")
    provider       = data.get("provider", "gmail")
    sender         = data.get("sender_email", "")
    password       = data.get("sender_pass", "")
    custom_host    = data.get("custom_host", "")
    custom_port    = data.get("custom_port", 587)
    use_template   = data.get("use_template", True)
    footer_text    = data.get("footer_text", "")
    personalize_on = data.get("personalize", False)

    if not recipients or not subject or not body or not sender or not password:
        return jsonify({"error": "Missing required fields"}), 400

    job_id = str(uuid.uuid4())
    job_store[job_id] = {"sent": 0, "total": len(recipients), "failed": [], "done": False, "cancelled": False}

    t = threading.Thread(
        target=run_send_job,
        args=(job_id, recipients, subject, body, provider, sender, password,
              use_template, footer_text, personalize_on, custom_host, custom_port),
        daemon=True
    )
    t.start()

    return jsonify({"job_id": job_id, "total": len(recipients)})

@app.route("/progress/<job_id>")
def progress(job_id):
    def stream():
        while True:
            job = job_store.get(job_id)
            if not job:
                yield f"data: {json.dumps({'error': 'Job not found'})}\n\n"
                break
            payload = {
                "sent":      job["sent"],
                "total":     job["total"],
                "failed":    job["failed"],
                "done":      job["done"],
                "cancelled": job.get("cancelled", False),
                "error":     job.get("error", ""),
            }
            yield f"data: {json.dumps(payload)}\n\n"
            if job["done"]:
                job_store.pop(job_id, None)
                break
            time.sleep(0.5)
    return Response(stream(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

@app.route("/cancel/<job_id>", methods=["POST"])
def cancel(job_id):
    job = job_store.get(job_id)
    if job:
        job["cancelled"] = True
        return jsonify({"ok": True})
    return jsonify({"error": "Job not found"}), 404

if __name__ == "__main__":
    app.run(debug=os.getenv("FLASK_DEBUG", "false").lower() == "true")
